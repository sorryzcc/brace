import re
from openpyxl import load_workbook

# 加载工作簿
wb = load_workbook('266总表0923.xlsx')
ws = wb.active  # 假设我们只处理活动表

# 正则表达式匹配并移除形如 {xxx} 的部分
pattern = r'\{.*?\}'

# 遍历D列，并在E列填入去除形如 {xxx} 后的内容
for row in ws.iter_rows(min_row=1, min_col=4, max_col=4):  # D列为第4列
    d_cell = row[0]
    if isinstance(d_cell.value, str):  # 确保值是字符串类型
        # 使用正则表达式移除形如 {xxx} 的部分
        new_value = re.sub(pattern, '', d_cell.value)
        
        # 在E列写入新值
        e_cell = ws.cell(row=d_cell.row, column=5)  # E列为第5列
        e_cell.value = new_value

# 保存工作簿
wb.save('unbraceTranslate.xlsx')