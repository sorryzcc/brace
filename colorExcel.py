from openpyxl import Workbook  
from openpyxl.styles import PatternFill  
  
# 准备数据  
data = [  
    ['Header1', 'Header2', 'Header3'],  
    ['Data1', 'Data2', 'HighlightThis'],  
    ['Data3', 'Data4', 'AnotherHighlight'],  
    ['Data5', 'Data6', 'Data7']  
]  
  
# 创建工作簿和工作表  
wb = Workbook()  
ws = wb.active  
ws.title = "Sheet1"  
  
# 填充数据  
for row in data:  
    ws.append(row)  
  
# 设置需要高亮颜色的单元格地址（这里以C列的值为'HighlightThis'和'AnotherHighlight'的单元格为例）  
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3):  # C列的列索引为3（因为是从1开始的）  
    for cell in row:  
        if cell.value == 'HighlightThis' or cell.value == 'AnotherHighlight':  
            # 设置填充颜色为黄色  
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
            cell.fill = fill  
  
# 保存Excel文件  
wb.save('output_with_colors.xlsx')  
  
print('Excel file with colors has been generated.')